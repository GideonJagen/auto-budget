import os
from dataclasses import dataclass
import openpyxl
from openpyxl.descriptors.base import String
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import xlrd
import pandas as pd
from pathlib import Path
from operator import itemgetter
from datetime import date

@dataclass
class Style():
    FONT_BIG_BOLD = Font(name='Arial', bold=True, size=13)
    FONT_SMALL_BOLD = Font(name='Arial', bold=True, size=11)
    FONT_STANDARD = Font(name='Arial')
    COLOR_YELLOW_1 = PatternFill(fgColor='e0d8c1', fill_type='solid')
    COLOR_YELLOW_2 = PatternFill(fgColor='e6e1d2', fill_type='solid')
    COLOR_YELLOW_3 = PatternFill(fgColor='eceae4', fill_type='solid')
    COLOR_BLUE_1 = PatternFill(fgColor='c1c9e0', fill_type='solid')
    COLOR_BLUE_2 = PatternFill(fgColor='d2d7e6', fill_type='solid')
    COLOR_BLUE_3 = PatternFill(fgColor='e4e6ec', fill_type='solid')

    BORDER_DOTTED = Border(
        left=Side(border_style='dotted', color='000000'), 
        top=Side(border_style='dotted', color='000000'),
        )

class AutoBudget:
    def __init__(self, input_dir_path) -> None:
        # Init
        self.workbook = openpyxl.Workbook()  # Create workbook
        self.cost_center_list = []
        self.budget_dict = self.load_budgets(input_dir_path)
        self.cost_types = {}

        self.month_header = [
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec",
        ]
        self.column_standard_header = ["Month", "Budget", "Diff"]
        self.offset = 2  # Offset from 0 where table begins
        self.year = ""

    def get_cost_centers(self) -> String:
        cost_centers = ""
        for cost_center in self.cost_center_list:
            cost_centers += cost_center + " "
        return cost_centers

    # ------------------------------------------------------------------------------------------------------------
    #           Load Data
    # ------------------------------------------------------------------------------------------------------------

    def load_budgets(self, input_dir_path) -> dict:
        budget_dict = {}
        doublet_check_list = []
        for file in os.listdir(input_dir_path):
            if file.endswith(".XLS"):
                file_path = os.path.join(input_dir_path, file)
                wb = xlrd.open_workbook(Path(file_path))
                cost_report = wb["Cost center report"]

                # Find cost_date, cost_center and table index
                for index, row in enumerate(cost_report.get_rows()):
                    info_cost_elemnt = row[5].value
                    if info_cost_elemnt == "Fiscal period / year (Interval, Req.)":
                        cost_date = row[6].value
                    elif info_cost_elemnt == "Cost Center Node":
                        cost_center = row[6].value[-4:]
                    elif info_cost_elemnt == "Table":
                        start_of_table = index + 1
                    elif info_cost_elemnt == "HSQVBI_CCTR_GR":
                        end_of_table = index

                # Make table array
                table = [
                    cost_report.row_values(row)
                    for row in range(start_of_table, end_of_table)
                ]

                # Make dataframe
                columns = table[0]
                columns[6] = "Cost Type"
                table_df = pd.DataFrame(table[1:], columns=table[0])
                table_df = table_df.set_index("Cost Type")

                # Remove unwanted columns and rows
                table_df = table_df.drop(columns=["Cost Element", ""])
                table_df = table_df.drop("")

                if cost_date not in budget_dict:
                    budget_dict[cost_date] = []

                check = cost_center + cost_date
                if check in doublet_check_list:
                    raise Exception(
                        f"cost center {cost_center} has a dublicate with the date {cost_date}, please remove it!"
                    )
                budget_dict[cost_date].append(
                    {cost_center: table_df, "id": cost_center}
                )
                doublet_check_list.append(check)
                if cost_center not in self.cost_center_list:
                    self.cost_center_list.append(cost_center)

        self.cost_center_list = sorted(self.cost_center_list)

        return budget_dict

    # ------------------------------------------------------------------------------------------------------------
    #           Write Data
    # ------------------------------------------------------------------------------------------------------------

    def write_to_cell(self, sheet, row, col, value, font, style=False) -> None:
        sheet.cell(row, col, value).font = font
        if style:
            sheet.cell(row, col).style = "Comma [0]"
            sheet.cell(row, col).alignment = Alignment(horizontal="right")

    # Sum different costs based on actual cost
    # pos: cost = 0 budget = 1
    def sum_month(self, cost_center_dict_list, pos) -> dict:
        month_dict = {}
        for dict in cost_center_dict_list:
            df = list(dict.values())[0]
            for index, row in df.iterrows():
                cost = row[pos]
                if not cost:
                    cost = 0
                if index not in month_dict:
                    month_dict[index] = float(cost)
                else:
                    month_dict[index] += float(cost)
        return month_dict

    def make_compilation(self) -> None:
        compilation_sheet = self.workbook.active
        compilation_sheet.title = "Summary Sheet"
        same_every_col = len(self.cost_center_list) + len(self.column_standard_header)

        # Column headers
        self.add_column_headers(compilation_sheet, same_every_col)

        # Add costs and cost types to worksheet
        for cost_date, cost_center_dict_list in self.budget_dict.items():
            i_col = (int(cost_date[:3]) - 1) * same_every_col + self.offset + 1
            self.year = cost_date[3:]

            # Actual & cost types
            month_cost_dict_actual = self.sum_month(cost_center_dict_list, 0)
            for cost_type, cost in month_cost_dict_actual.items():
                if cost_type in self.cost_types.keys():
                    self.write_to_cell(compilation_sheet, self.cost_types.get(cost_type), i_col, cost, Style.FONT_STANDARD, style=True)
                else:
                    i_row = compilation_sheet.max_row + 1
                    self.write_to_cell(compilation_sheet, i_row, self.offset, cost_type, Style.FONT_STANDARD) # Add cost type
                    self.cost_types[cost_type] = i_row
                    self.write_to_cell(compilation_sheet, i_row, i_col, cost, Style.FONT_STANDARD, style=True)


            # Planned
            month_cost_dict_planned = self.sum_month(cost_center_dict_list, 1)
            for cost_type, cost in month_cost_dict_planned.items():
                self.write_to_cell(compilation_sheet, self.cost_types.get(cost_type), i_col+1, cost, Style.FONT_STANDARD, style=True)


            # Add cost for individual cost centers
            cost_center_dict_list = sorted(cost_center_dict_list, key=itemgetter("id"))
            for i, cost_center_dict in enumerate(cost_center_dict_list):
                cost_center = cost_center_dict["id"]
                offset_individual = len(
                    self.column_standard_header
                ) + self.cost_center_list.index(cost_center)
                i_col_individual = i_col + offset_individual

                # Check that cost is ending up in the right place
                header_value = compilation_sheet.cell(
                    self.offset, i_col_individual
                ).value
                if not cost_center == header_value:
                    raise Exception(
                        f"The cost center in header is {header_value}. The cost center for the data is {cost_center} for date {cost_date}"
                    )

                # Add cost for all cost types
                for cost_type, row in cost_center_dict[cost_center].iterrows():
                    actual_cost = row[0]
                    if actual_cost:
                        self.write_to_cell(compilation_sheet, self.cost_types.get(cost_type), i_col_individual, actual_cost, Style.FONT_STANDARD, style=True)

        # Fill in blank cells
        months_with_data = len(self.budget_dict)
        for i_col in range(
            self.offset + 1, self.offset + 1 + months_with_data * same_every_col
        ):
            for i_row in range(self.offset + 1, compilation_sheet.max_row + 1):
                if not compilation_sheet.cell(i_row, i_col).value:
                    self.write_to_cell(compilation_sheet, i_row, i_col, 0, Style.FONT_STANDARD, style=True)

        # Differential Actual-Planned
        col = self.offset + 3
        for i_col in range(col, compilation_sheet.max_column+1, same_every_col):
            for i_row in range(self.offset +1, compilation_sheet.max_row+1):
                budget_col_letter = get_column_letter(i_col-1)
                actual_col_letter = get_column_letter(i_col-2)
                self.write_to_cell(compilation_sheet, i_row, i_col, f"={budget_col_letter}{i_row}-{actual_col_letter}{i_row}", Style.FONT_STANDARD, style=True)
        

        # Add a sum of months for each row
        month_columns = range(
            self.offset + 1, compilation_sheet.max_column, same_every_col
        )
        for row in range(self.offset + 1, compilation_sheet.max_row + 1):
            cell_value = f"="
            for month_col in month_columns:
                cell_value += f"+ {get_column_letter(month_col)}{row}"
            self.write_to_cell(compilation_sheet, row, compilation_sheet.max_column, cell_value, Style.FONT_SMALL_BOLD, style=True)


        self.make_sum_rows(compilation_sheet, same_every_col)
        self.style_sheet(compilation_sheet, same_every_col)

    def add_column_headers(self, sheet, same_every_col) -> None:
        # Add title to table
        #self.write_to_cell(sheet, self.offset, self.offset, sheet.title, Style.FONT_BIG_BOLD)

        # Add standard headers to worksheet
        for i in range(len(self.month_header)):
            i_col =  i*same_every_col+self.offset+1
            self.write_to_cell(sheet, self.offset, i_col, self.month_header[i], Style.FONT_BIG_BOLD)
            i_col +=1
            self.write_to_cell(sheet, self.offset, i_col, self.column_standard_header[1], Style.FONT_SMALL_BOLD)
            i_col +=1
            self.write_to_cell(sheet, self.offset, i_col, self.column_standard_header[2], Style.FONT_SMALL_BOLD)
            for cost_center in sorted(self.cost_center_list):
                i_col +=1
                self.write_to_cell(sheet, self.offset, i_col, cost_center, Style.FONT_SMALL_BOLD)
        
        # Add Sum title in end
        self.write_to_cell(sheet, self.offset, sheet.max_column + 1, "Sum:", Style.FONT_BIG_BOLD, style=False)

        # Add Sum title in end
        self.write_to_cell(
            sheet,
            self.offset,
            sheet.max_column + 1,
            "Sum:",
            self.font_bold,
            style=False,
        )

    def make_sum_rows(self, sheet, same_every_col) -> None:
        # Sum all columns
        row_total = sheet.max_row+2
        self.write_to_cell(sheet, row_total, self.offset, "Cost", Style.FONT_SMALL_BOLD)
        for col in range(self.offset+1,sheet.max_column):
            column_letter = get_column_letter(col)
            self.write_to_cell(sheet, row_total, col, f"=SUM({column_letter}{2}:{column_letter}{row_total-2})", Style.FONT_SMALL_BOLD, style=True)
        
        # Move budget sums
        row = sheet.max_row+1
        self.write_to_cell(sheet, row, self.offset, "Budget", Style.FONT_SMALL_BOLD)
        for col in range(self.offset+2, sheet.max_column, same_every_col):
            budget_sum = sheet.cell(row_total, col).value
            self.write_to_cell(sheet, row_total, col, "-", Style.FONT_STANDARD, style=True)
            if ('00'+str(int((col-self.offset)/same_every_col+1))+self.year) in self.budget_dict.keys():
                self.write_to_cell(sheet, row, col-1, budget_sum, Style.FONT_SMALL_BOLD, style=True)
            else:
                self.write_to_cell(sheet, row, col-1, f"=MEDIAN({get_column_letter(self.offset+1)}{row}:{get_column_letter(col-2)}{row})", Style.FONT_SMALL_BOLD, style=True)

        # Accumulation of sums
        row = sheet.max_row+1
        self.write_to_cell(sheet, row, self.offset, "Cost (ACC)", Style.FONT_SMALL_BOLD)
        for col in range(self.offset+1, sheet.max_column, same_every_col):
            column_letter = get_column_letter(col)
            if col == self.offset+1:
                self.write_to_cell(sheet, row, col, f"=SUM({column_letter}{row-2}+0)", Style.FONT_SMALL_BOLD, style=True)
            else:
                self.write_to_cell(sheet, row, col, f"=SUM({column_letter}{row-2}+{get_column_letter(col-same_every_col)}{row})", Style.FONT_SMALL_BOLD, style=True)
        
        # Accumulation of budgets
        row = sheet.max_row+1
        self.write_to_cell(sheet, row, self.offset, "Budget (ACC)", Style.FONT_SMALL_BOLD)
        for col in range(self.offset+1, sheet.max_column, same_every_col):
            column_letter = get_column_letter(col)
            if col == self.offset+1:
                self.write_to_cell(sheet, row, col, f"=SUM({column_letter}{row-2}+0)", Style.FONT_SMALL_BOLD, style=True)
            else:
                self.write_to_cell(sheet, row, col, f"=SUM({column_letter}{row-2}+{get_column_letter(col-same_every_col)}{row})", Style.FONT_SMALL_BOLD, style=True)
        
        # Differential row
        row = sheet.max_row+1
        self.write_to_cell(sheet, row, self.offset, "Diff", Style.FONT_SMALL_BOLD)
        for col in range(self.offset+1, sheet.max_column, same_every_col):
            column_letter = get_column_letter(col)
            self.write_to_cell(sheet, row, col, f"={column_letter}{row-3} - {column_letter}{row-4}", Style.FONT_SMALL_BOLD, style=True)

            # Delete total duplicate in diff column
            self.write_to_cell(sheet, row_total, col + 2, "-", Style.FONT_STANDARD, style=True)

        # Differential (ACC) row 
        row = sheet.max_row+1
        self.write_to_cell(sheet, row, self.offset, "Diff (ACC)", Style.FONT_SMALL_BOLD)
        for col in range(self.offset+1, sheet.max_column, same_every_col):
            column_letter = get_column_letter(col)
            self.write_to_cell(sheet, row, col, f"={column_letter}{row-2} - {column_letter}{row-3}", Style.FONT_SMALL_BOLD, style=True)
                

    # ------------------------------------------------------------------------------------------------------------
    #           Fix style
    # ------------------------------------------------------------------------------------------------------------

    def autosize_column(self, ws, columnrange, length=0) -> None:
        for column in columnrange:
            column_cells = [c for c in ws.columns][column - 1]
            if not length:
                length = max(
                    len(str(cell.value)) * 1.05 for cell in column_cells
                )  # Libre Office: *0.87, Microsoft Excel: *1.05
            ws.column_dimensions[column_cells[0].column_letter].width = length

    # Set thich border around given range
    def set_thick_border(self, sheet, startRow, startCol, endRow, endCol):
        max_y = endRow - startRow  # index of the last row
        for pos_y, r in enumerate(range(startRow, endRow + 1)):
            max_x = endCol - startCol  # index of the last cell
            for pos_x, c in enumerate(range(startCol, endCol + 1)):
                cell = sheet.cell(r, c)
                BORDER = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )
                if pos_x == 0:
                    BORDER.left = Side(border_style="thick", color="000000")
                if pos_x == max_x:
                    BORDER.right = Side(border_style="thick", color="000000")
                if pos_y == 0:
                    BORDER.top = Side(border_style="thick", color="000000")
                if pos_y == max_y:
                    BORDER.bottom = Side(border_style="thick", color="000000")
                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = BORDER

    def style_sheet(self, sheet, same_every_col) -> None:
        # Set column colors
        color_1 = [Style.COLOR_BLUE_1, Style.COLOR_YELLOW_1]
        color_2 = [Style.COLOR_BLUE_2, Style.COLOR_YELLOW_2]
        color_3 = [Style.COLOR_BLUE_3, Style.COLOR_YELLOW_3]
        i_color = 1
        for i_col in range(self.offset, sheet.max_column):
            column_header = sheet.cell(self.offset, i_col).value
            if column_header in self.month_header:
                i_color += 1
            for i_row in range(self.offset, sheet.max_row + 1):
                if i_col == self.offset:
                    pass
                elif column_header in self.month_header:
                    sheet.cell(i_row, i_col).fill = color_1[i_color % 2]
                elif column_header in self.column_standard_header:
                    sheet.cell(i_row, i_col).fill = color_2[i_color % 2]
                else:
                    sheet.cell(i_row, i_col).fill = color_3[i_color % 2]

        # Set column borders
        for col in range(self.offset, sheet.max_column + 1):
            for row in range(self.offset, sheet.max_row + 1):
                sheet.cell(row, col).border = Style.BORDER_DOTTED

        # Upper rows
        self.set_thick_border(
            sheet,
            self.offset,
            self.offset,
            self.offset + len(self.cost_types.keys()),
            self.offset,
        )
        self.set_thick_border(
            sheet,
            self.offset,
            self.offset + 1,
            self.offset + len(self.cost_types.keys()),
            sheet.max_column - 1,
        )
        self.set_thick_border(
            sheet,
            self.offset,
            sheet.max_column,
            self.offset + len(self.cost_types.keys()),
            sheet.max_column,
        )
        # Lower rows
        self.set_thick_border(
            sheet,
            self.offset + len(self.cost_types.keys()) + 1,
            self.offset,
            sheet.max_row,
            self.offset,
        )
        self.set_thick_border(
            sheet,
            self.offset + len(self.cost_types.keys()) + 1,
            self.offset + 1,
            sheet.max_row,
            sheet.max_column - 1,
        )
        self.set_thick_border(
            sheet,
            self.offset + len(self.cost_types.keys()) + 1,
            sheet.max_column,
            sheet.max_row,
            sheet.max_column,
        )

        # Hide columns
        sheet.sheet_properties.outlinePr.summaryRight = False
        for i in range(
            same_every_col - 1
        ):  # Grouping several at once does not work, but one at a time works
            for col in range(self.offset + 2 + i, sheet.max_column, same_every_col):
                sheet.column_dimensions.group(
                    get_column_letter(col), get_column_letter(col), hidden=True
                )

        # Size columns
        self.autosize_column(sheet, [self.offset])
        self.autosize_column(sheet, range(self.offset + 1, sheet.max_column + 1), 11)


if __name__ == "__main__":

    print("Hello, I am your budget automator")
    # budget = AutoBudget("./dummydata/") # If you want to run with dummydata
    budget = AutoBudget("./data/")  # If you want to run with data
    budget.make_compilation()
    budget.workbook.save(
        f"Cost Report Summary {date.today()} ({budget.get_cost_centers()}).xlsx"
    )
